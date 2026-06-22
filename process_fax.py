"""FAX Order Processing Script - Main entry point

Usage:
    python process_fax.py <pdf_path> [--staff NAME] [--remarks TEXT]
    python process_fax.py <pdf_path1> <pdf_path2> ...  (multiple files)

Output:
    - CSV file with order data (output/ directory)
    - Sylvia order PDF (if applicable)
    - Haruna order PDF (if applicable)
"""
import os
import sys
import csv
import argparse

# Windows terminal UTF-8 output
if sys.stdout.encoding and sys.stdout.encoding.lower() in ('cp932', 'shift_jis', 'mbcs'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from datetime import date

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from ocr_module import process_fax_pdf, load_staff
from pdf_generator import gen_sylvia_pdf, gen_haruna_pdf, gen_nomura_pdf, gen_emsink_pdf, gen_powbar_pdf

# パートナー別 PDF生成器マップ (output_dest -> (関数, ファイル名プレフィックス))
PARTNER_PDF_GENERATORS = {
    "野村不動産": (gen_nomura_pdf, "野村不動産"),
    "エムズインク": (gen_emsink_pdf, "エムズインク"),
    "POW BAR": (gen_powbar_pdf, "POWBAR"),
}


OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
HISTORY_DIR = os.path.join(OUTPUT_DIR, "履歴")


def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(HISTORY_DIR, exist_ok=True)


def _archive_existing_file(path: str) -> str | None:
    """書込先に既存ファイルがあれば、履歴フォルダへ退避する。

    退避先ファイル名: `<basename>_YYYYMMDDHHMMSS.<ext>`

    Args:
        path: これから書き込もうとするファイルのフルパス

    Returns:
        退避したファイルのフルパス。退避不要の場合は None。
    """
    if not os.path.exists(path):
        return None
    try:
        os.makedirs(HISTORY_DIR, exist_ok=True)
        base = os.path.basename(path)
        name, ext = os.path.splitext(base)
        from datetime import datetime as _dt
        ts = _dt.now().strftime("%Y%m%d%H%M%S")
        archived = os.path.join(HISTORY_DIR, f"{name}_{ts}{ext}")
        # 衝突回避（同秒に複数回呼ばれた場合）
        if os.path.exists(archived):
            archived = os.path.join(HISTORY_DIR, f"{name}_{ts}_{os.getpid()}{ext}")
        import shutil as _shutil
        _shutil.move(path, archived)
        return archived
    except OSError as e:
        # 退避失敗は致命的ではない（書き込みは続行）
        print(f"[output履歴] 退避失敗 {path}: {e}")
        return None


def parse_infomart_csv(csv_bytes, filename="infomart.csv"):
    """インフォマートCSVを読み込み、OCR結果と同じresults形式に変換する。

    フォーマット:
    - Row 0: H行（ヘッダー、日付）
    - Row 1: カラム名（［データ区分］等）
    - Row 2+: D行（データ）
    - 最終行: F行（フッター）
    - 伝票Noでグループ化
    """
    import unicodedata
    from ocr_module import match_product, match_ddc, load_product_master, load_ddc_master

    pm = load_product_master()
    ddc_master = load_ddc_master()

    text = None
    for enc in ['shift-jis', 'cp932', 'utf-8-sig', 'utf-8']:
        try:
            text = csv_bytes.decode(enc)
            break
        except (UnicodeDecodeError, AttributeError):
            continue
    if text is None:
        return []

    import io
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)

    # ヘッダー行の位置を自動判定（取引先により H行 ありなしのフォーマット差異あり）
    # Pattern A (H行あり): Row0=H行(日付) / Row1=カラム名 / Row2+=D行
    # Pattern B (H行なし): Row0=カラム名 / Row1+=D行
    if not all_rows:
        return []
    header_row_idx = None
    for i, r in enumerate(all_rows[:3]):
        if r and r[0].strip().strip('［］').strip('[]') == 'データ区分':
            header_row_idx = i
            break
    if header_row_idx is None:
        return []

    headers = [h.strip('［］').strip('[]') for h in all_rows[header_row_idx]]
    data_rows = []
    for r in all_rows[header_row_idx + 1:]:
        if r and r[0].strip() == 'D':
            row_dict = {}
            for j, h in enumerate(headers):
                row_dict[h] = r[j] if j < len(r) else ''
            data_rows.append(row_dict)

    if not data_rows:
        return []

    # 伝票Noでグループ化
    from collections import OrderedDict
    slip_groups = OrderedDict()
    for r in data_rows:
        slip_no = r.get('伝票No', '').strip()
        if slip_no not in slip_groups:
            slip_groups[slip_no] = []
        slip_groups[slip_no].append(r)

    results = []
    page = 0
    for slip_no, group_rows in slip_groups.items():
        page += 1
        first = group_rows[0]

        # 取引先名・納品場所から納品先を判定
        dest_name = first.get('納品場所名', '').strip()
        if not dest_name:
            dest_name = first.get('取引先名', '').strip()
        dest_name = unicodedata.normalize('NFKC', dest_name)
        dest_address = unicodedata.normalize('NFKC', first.get('納品場所 住所', '').strip())
        sender = unicodedata.normalize('NFKC', first.get('取引先名', '').strip())
        delivery_date = first.get('納品日', '').strip()

        # DDCマッチング
        ddc_match = match_ddc(dest_name, ddc_master, sender=sender)
        # 住所をDDCマッチ結果に補完（マスタにない場合）
        if ddc_match.get("matched") and not ddc_match.get("address") and dest_address:
            ddc_match["address"] = dest_address

        # 商品マッチング
        matched_items = []
        for r in group_rows:
            product_code = r.get('自社管理商品コード', '').strip()
            product_name = unicodedata.normalize('NFKC', r.get('商品名', '').strip())
            try:
                qty = float(r.get('数量', '0').strip() or '0')
                qty = int(qty) if qty == int(qty) else qty
            except ValueError:
                qty = 0

            ocr_item = {
                "jan_code": "",
                "product_code": product_code,
                "product_name": product_name,
                "quantity_cs": qty,
            }
            match = match_product(ocr_item, pm)
            if match:
                matched_items.append(match)
            else:
                matched_items.append({
                    "matched": False,
                    "ocr_name": product_name,
                    "jan": "",
                    "code": product_code,
                    "quantity": qty,
                })

        sylvia_items = [i for i in matched_items if i.get("matched") and i.get("output_dest") == "シルビア"]
        haruna_items = [i for i in matched_items if i.get("matched") and i.get("output_dest") == "ハルナ"]

        results.append({
            "page": page,
            "source": "infomart",
            "ocr_raw": {
                "order_no": slip_no,
                "delivery_date": delivery_date.replace("/", "-"),
                "delivery_dest": dest_name,
                "sender": sender,
                "notes": "",
                "items": [],
            },
            "matched_items": matched_items,
            "ddc_match": ddc_match,
            "sylvia_items": sylvia_items,
            "haruna_items": haruna_items,
        })

    return results


def parse_smacla_csv(csv_bytes, filename="smacla.csv"):
    """スマクラ受注伝票CSV（株式会社スタイリングライフ・ホールディングス 等）を読み込み、results形式に変換する。

    特徴:
    - 1メッセージID毎に1受注（複数商品=複数行で構成）
    - エンコーディング: Shift-JIS
    - 数量は「発注数量（バラ）」=本数で記載されているので、入数でCS数に換算
    - JANコードで商品マッチング、納品先名でDDCマッチング
    """
    import unicodedata
    import io
    from collections import OrderedDict
    from ocr_module import match_product, match_ddc, load_product_master, load_ddc_master

    pm = load_product_master()
    ddc_master = load_ddc_master()

    # CSV読込（Shift-JIS優先）
    text = None
    for enc in ['shift_jis', 'cp932', 'utf-8-sig', 'utf-8']:
        try:
            text = csv_bytes.decode(enc)
            break
        except (UnicodeDecodeError, AttributeError):
            continue
    if text is None:
        return []

    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []

    # メッセージ識別IDでグループ化（1メッセージ=1受注）
    msg_groups = OrderedDict()
    for r in rows:
        msg_id = (r.get('メッセージ識別ID') or '').strip()
        if not msg_id:
            continue
        msg_groups.setdefault(msg_id, []).append(r)

    results = []
    page = 0
    for msg_id, group_rows in msg_groups.items():
        page += 1
        first = group_rows[0]

        # ヘッダー情報
        sender = unicodedata.normalize('NFKC', (first.get('発注者名称') or '').strip())
        dest_name = unicodedata.normalize('NFKC', (first.get('最終納品先名称') or '').strip())
        nohinsaki_code = (first.get('最終納品先コード') or '').strip()
        # 取引番号（顧客側の発注書番号）を伝票番号として使う
        slip_no = (first.get('取引番号（発注・返品）') or '').strip()
        order_date = (first.get('発注日') or '').strip()
        delivery_date = (first.get('最終納品先納品日') or first.get('直接納品先納品日') or '').strip()

        # 日付フォーマット変換: YYYYMMDD → YYYY-MM-DD
        def _fmt_date(s):
            s = (s or '').strip()
            if len(s) == 8 and s.isdigit():
                return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"
            return s

        order_date = _fmt_date(order_date)
        delivery_date = _fmt_date(delivery_date)

        # DDCマッチング（コード優先→名前）
        ddc_match = None
        if nohinsaki_code and ddc_master is not None and '納品先コード' in ddc_master.columns:
            mask = ddc_master['納品先コード'].astype(str).str.strip() == nohinsaki_code
            if mask.any():
                from ocr_module import _ddc_row_to_dict
                ddc_match = _ddc_row_to_dict(ddc_master[mask].iloc[0])
                ddc_match['match_score'] = 1.0
                ddc_match['low_confidence'] = False
                ddc_match['candidates'] = []
        if ddc_match is None:
            ddc_match = match_ddc(dest_name, ddc_master, sender=sender)

        # 商品行
        matched_items = []
        for r in group_rows:
            jan = (r.get('商品コード（ＧTIN）') or r.get('商品コード（発注用）') or '').strip()
            product_name = unicodedata.normalize('NFKC', (r.get('商品名') or '').strip())
            try:
                bara_qty = int(r.get('発注数量（バラ）') or 0)
            except (ValueError, TypeError):
                bara_qty = 0
            try:
                unit_price = float(r.get('売単価') or 0)
            except (ValueError, TypeError):
                unit_price = 0.0
            try:
                amount = float(r.get('売価金額') or 0)
            except (ValueError, TypeError):
                amount = 0.0

            # 商品マスタから入数取得 → CS換算
            cs_qty = 0
            irisuu = 0
            if pm is not None and 'JANコード' in pm.columns:
                pmask = pm['JANコード'].astype(str).str.strip() == jan
                if pmask.any():
                    prow = pm[pmask].iloc[0]
                    try:
                        irisuu = int(prow.get('入数') or 0)
                    except (ValueError, TypeError):
                        irisuu = 0
            if irisuu > 0 and bara_qty > 0:
                # 整数CS換算（バラが入数の倍数の前提）
                cs_qty = bara_qty // irisuu

            ocr_item = {
                'jan_code': jan,
                'product_name': product_name,
                'quantity_cs': cs_qty,
            }
            match = match_product(ocr_item, pm)
            if match:
                # スマクラ側の単価情報を反映（スマクラの売単価＝当方の卸単価）
                if unit_price:
                    match['unit_price'] = unit_price
                if amount:
                    match['amount'] = amount
                matched_items.append(match)
            else:
                matched_items.append({
                    'matched': False,
                    'ocr_name': product_name,
                    'jan': jan,
                    'quantity': cs_qty,
                    'unit_price': unit_price,
                    'amount': amount,
                })

        # 出力先でグルーピング
        sylvia_items = [i for i in matched_items if i.get('matched') and i.get('output_dest') == 'シルビア']
        haruna_items = [i for i in matched_items if i.get('matched') and i.get('output_dest') == 'ハルナ']

        results.append({
            'page': page,
            'source': 'smacla',
            'source_file': filename,
            'ocr_raw': {
                'order_no': slip_no or msg_id,
                'order_date': order_date,
                'delivery_date': delivery_date,
                'delivery_dest': dest_name,
                'sender': sender,
                'notes': f'スマクラ受注 / メッセージID:{msg_id}',
                'items': [],
            },
            'matched_items': matched_items,
            'ddc_match': ddc_match,
            'sylvia_items': sylvia_items,
            'haruna_items': haruna_items,
        })

    return results


def parse_paltac_csv(csv_bytes, filename="paltac.csv"):
    """PALTAC WebEDI CSVを読み込み、OCR結果と同じresults形式に変換する。

    - 届先区分「発注」のみ取り込み（返品は除外）
    - 同一伝票番号の複数行を1ページにまとめる
    - JANコードで商品マッチング
    - 届先名でDDCマッチング
    """
    import unicodedata
    from ocr_module import match_product, match_ddc, load_product_master, load_ddc_master, normalize

    pm = load_product_master()
    ddc_master = load_ddc_master()

    # CSVを読み込み（Shift-JIS）
    text = None
    for enc in ['shift-jis', 'cp932', 'utf-8-sig', 'utf-8']:
        try:
            text = csv_bytes.decode(enc)
            break
        except (UnicodeDecodeError, AttributeError):
            continue
    if text is None:
        return []

    import io
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)

    # 届先区分「発注」のみフィルタ
    order_rows = [r for r in rows if unicodedata.normalize('NFKC', r.get('届先区分', '').strip()) == '発注']
    if not order_rows:
        return []

    # 伝票番号でグループ化
    from collections import OrderedDict
    slip_groups = OrderedDict()
    for r in order_rows:
        slip_no = r.get('伝票番号', '').strip()
        if slip_no not in slip_groups:
            slip_groups[slip_no] = []
        slip_groups[slip_no].append(r)

    results = []
    page = 0
    for slip_no, group_rows in slip_groups.items():
        page += 1
        first = group_rows[0]

        # 届先名を正規化
        dest_name = unicodedata.normalize('NFKC', first.get('届先名', '').strip())
        sender = unicodedata.normalize('NFKC', first.get('発注元名', '').strip())
        order_date = first.get('発注日', '').strip()
        delivery_date = first.get('着荷指定日', '').strip()

        # DDCマッチング
        ddc_match = match_ddc(dest_name, ddc_master, sender=sender)

        # 商品マッチング
        matched_items = []
        for r in group_rows:
            jan = r.get('商品ｺｰﾄﾞ', '').strip()
            product_name = unicodedata.normalize('NFKC', r.get('商品名', '').strip())
            try:
                cs_qty = int(r.get('ｹｰｽ数', '0').strip() or '0')
            except ValueError:
                cs_qty = 0
            try:
                bara_qty = int(r.get('ﾊﾞﾗ数', '0').strip() or '0')
            except ValueError:
                bara_qty = 0
            try:
                irisuu = int(r.get('入数', '0').strip() or '0')
            except ValueError:
                irisuu = 0

            # CS数が0でバラ数がある場合、CS換算
            if cs_qty == 0 and bara_qty > 0 and irisuu > 0:
                cs_qty = bara_qty / irisuu  # 端数はそのまま

            ocr_item = {
                "jan_code": jan,
                "product_name": product_name,
                "quantity_cs": cs_qty,
            }
            match = match_product(ocr_item, pm)
            if match:
                matched_items.append(match)
            else:
                matched_items.append({
                    "matched": False,
                    "ocr_name": product_name,
                    "jan": jan,
                    "quantity": int(cs_qty) if cs_qty == int(cs_qty) else cs_qty,
                })

        # 出力先で分類
        sylvia_items = [i for i in matched_items if i.get("matched") and i.get("output_dest") == "シルビア"]
        haruna_items = [i for i in matched_items if i.get("matched") and i.get("output_dest") == "ハルナ"]

        results.append({
            "page": page,
            "source": "paltac",
            "ocr_raw": {
                "order_no": slip_no,
                "delivery_date": delivery_date.replace("/", "-"),
                "delivery_dest": dest_name,
                "sender": sender,
                "notes": "",
                "items": [],
            },
            "matched_items": matched_items,
            "ddc_match": ddc_match,
            "sylvia_items": sylvia_items,
            "haruna_items": haruna_items,
        })

    return results


def results_to_csv(results, pdf_name):
    """Write OCR+matching results to CSV"""
    ensure_output_dir()
    csv_name = os.path.splitext(pdf_name)[0] + "_result.csv"
    csv_path = os.path.join(OUTPUT_DIR, csv_name)

    rows = []
    for page_result in results:
        if "error" in page_result:
            continue
        ocr = page_result["ocr_raw"]
        ddc = page_result["ddc_match"]
        candidates = ddc.get("candidates", [])

        # 候補列（最大3件）
        def cand_cell(idx):
            if idx < len(candidates):
                c = candidates[idx]
                return f"{c['name']} ({c['score']:.0%})"
            return ""

        low_conf = ddc.get("low_confidence", False)
        match_status = "OK" if ddc.get("matched") else "NG"
        if ddc.get("matched") and low_conf:
            match_status = "要確認"

        for item in page_result["matched_items"]:
            rows.append({
                "ページ": page_result["page"],
                "オーダーNO": ocr.get("order_no", ""),
                "納品日": ocr.get("delivery_date", ""),
                "発注元": ocr.get("sender", ""),
                "納品先(OCR)": ocr.get("delivery_dest", ""),
                "納品先(マスタ)": ddc.get("name", "") if ddc.get("matched") else "",
                "納品先住所": ddc.get("address", "") if ddc.get("matched") else "",
                "納品先TEL": ddc.get("tel", "") if ddc.get("matched") else "",
                "DDCマッチ": match_status,
                "DDC候補1": cand_cell(0),
                "DDC候補2": cand_cell(1),
                "DDC候補3": cand_cell(2),
                "商品名(OCR)": item.get("ocr_name", ""),
                "商品名(マスタ)": item.get("master_name", ""),
                "JANコード": item.get("jan", ""),
                "商品コード": item.get("code", ""),
                "規格": item.get("spec", ""),
                "配送荷姿": item.get("pack", ""),
                "CS単価": item.get("cs_price", ""),
                "数量(CS)": item.get("quantity", ""),
                "金額": item.get("amount", ""),
                "出力先": item.get("output_dest", ""),
                "商品マッチ": "OK" if item.get("matched") else "NG",
                "備考": ocr.get("notes", ""),
            })

    _archive_existing_file(csv_path)
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        if rows:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        else:
            f.write("データなし\n")

    return csv_path


# ── ネクストエンジン汎用受注CSV ──

NE_HEADERS = [
    "店舗伝票番号", "受注日", "受注郵便番号", "受注住所１", "受注住所２",
    "受注名", "受注名カナ", "受注電話番号", "受注メールアドレス",
    "発送郵便番号", "発送先住所１", "発送先住所２", "発送先名", "発送先カナ",
    "発送電話番号", "支払方法", "発送方法", "商品計", "税金", "発送料",
    "手数料", "ポイント", "その他費用", "合計金額", "ギフトフラグ",
    "時間帯指定", "日付指定", "作業者欄", "備考",
    "商品名", "商品コード", "商品価格", "受注数量", "商品オプション",
    "出荷済フラグ", "顧客区分", "顧客コード", "消費税率（%）",
    "のし", "ラッピング", "メッセージ",
]

# TWO会社情報（受注側の固定値）
_TWO_POSTAL = "1500012"
_TWO_ADDRESS = "東京都品川区西五反田2-24-4 THE CROSS GOTANDA 5階"
_TWO_NAME = "株式会社ＴＷＯ"
_TWO_TEL = "03-6839-0010"

def _get_irisuu(item):
    """商品マッチ結果から入数を取得する。case_quantity があればそれを使い、なければ1。"""
    cq = item.get("case_quantity", 0)
    try:
        cq = int(cq or 0)
    except (ValueError, TypeError):
        cq = 0
    return cq if cq > 0 else 1


def results_to_ne_csv(results, pdf_name):
    """OCR結果をネクストエンジン汎用受注CSV形式に変換して出力する。

    NE仕様:
    - 同一伝票番号の複数商品は行を分けて記載（ヘッダー部分は同じ値を繰り返す）
    - エンコーディング: Shift-JIS (cp932)
    - 店舗伝票番号 = オーダーNO
    - 発送先 = 納品先(DDCマッチ結果)
    - 受注側 = TWO固定情報
    """
    ensure_output_dir()
    csv_name = os.path.splitext(pdf_name)[0] + "_NE受注.csv"
    csv_path = os.path.join(OUTPUT_DIR, csv_name)

    rows = []
    for page_result in results:
        if "error" in page_result:
            continue
        ocr = page_result["ocr_raw"]
        ddc = page_result.get("ddc_match", {})

        order_no = ocr.get("order_no", "")
        delivery_date = ocr.get("delivery_date", "")
        # 受注日: YYYY/MM/DD 形式に変換
        order_date_str = ""
        if delivery_date:
            order_date_str = str(date.today()).replace("-", "/") + " 00:00:00"

        # 納品先情報
        dest_name = ddc.get("name", ocr.get("delivery_dest", ""))
        dest_postal = (ddc.get("postal", "") or "").replace("-", "")
        dest_address = ddc.get("address", "") or ""
        dest_tel = ddc.get("tel", "") or ""

        # 日付指定 = 納品日
        date_spec = ""
        if delivery_date:
            date_spec = delivery_date.replace("-", "/")

        # 商品行を生成
        matched_items = page_result.get("matched_items", [])
        if not matched_items:
            continue

        # 商品計・合計を計算（バラ単価 × バラ数量）
        total_amount = 0
        for item in matched_items:
            if item.get("matched"):
                irisuu = _get_irisuu(item)
                cs_p = item.get("cs_price", 0)
                qty = item.get("quantity", 0)
                try:
                    unit_p = int(float(cs_p or 0)) // irisuu if irisuu > 0 else int(float(cs_p or 0))
                    q = int(float(qty or 0)) * irisuu
                    total_amount += unit_p * q
                except (ValueError, TypeError):
                    pass

        tax = int(total_amount * 0.1)

        for item in matched_items:
            if not item.get("matched"):
                continue

            qty_cs = item.get("quantity", 0)
            cs_price = item.get("cs_price", 0)
            try:
                price = int(float(cs_price or 0))
            except (ValueError, TypeError):
                price = 0
            try:
                qty_cs_int = int(float(qty_cs or 0))
            except (ValueError, TypeError):
                qty_cs_int = 0

            jan = item.get("jan", "")
            # CS数 → バラ数に換算（case_quantityから動的取得）
            irisuu = _get_irisuu(item)
            quantity_bara = qty_cs_int * irisuu
            # バラ単価 = CS単価 / 入数
            unit_price_bara = int(price / irisuu) if irisuu > 0 else price

            product_name = item.get("master_name", item.get("ocr_name", ""))
            product_code = item.get("code", "")
            if not product_code:
                raise ValueError(f"商品コード未登録: {product_name} (JAN:{jan})")
            # 備考 = 直送のハルナ・シルビアのみ表記。それ以外（自社倉庫・ロット割れ自社倉庫経由）は空欄
            # shipping_route は web_app.py の確定処理時に各itemにセット済み
            # （無い場合のフォールバック: output_dest=="自社倉庫" でなく shipping_route 未設定なら direct 扱い）
            output_dest = item.get("output_dest", "")
            shipping_route = item.get("shipping_route", "")
            if not shipping_route:
                # フォールバック: shipping_route 未セット時
                shipping_route = "warehouse" if output_dest == "自社倉庫" else "direct"
            if shipping_route == "direct" and output_dest in (
                    "ハルナ", "シルビア", "野村不動産", "エムズインク", "POW BAR"):
                notes = output_dest
            else:
                notes = ""

            # 賞味期限を備考に追加（運用上は自社倉庫出荷時のみ入力される想定）
            expiry = (item.get("expiry_date") or "").strip()
            if expiry:
                expiry_text = f"賞味期限:{expiry.replace('-', '/')}"
                notes = f"{notes} {expiry_text}" if notes else expiry_text

            row = {
                "店舗伝票番号": order_no,
                "受注日": order_date_str,
                "受注郵便番号": dest_postal,
                "受注住所１": dest_address,
                "受注住所２": "",
                "受注名": dest_name,
                "受注名カナ": "",
                "受注電話番号": dest_tel,
                "受注メールアドレス": "",
                "発送郵便番号": dest_postal,
                "発送先住所１": dest_address,
                "発送先住所２": "",
                "発送先名": dest_name,
                "発送先カナ": "",
                "発送電話番号": dest_tel,
                "支払方法": "請求書払い",
                "発送方法": "",
                "商品計": total_amount,
                "税金": tax,
                # Phase 4: ロット割れ自社倉庫出荷時に shipping_fee_calculator で算出した送料を反映
                # （TWO負担の場合は 0、未計算の場合も 0）
                "発送料": int(page_result.get("shipping_fee") or 0),
                "手数料": 0,
                "ポイント": 0,
                "その他費用": 0,
                "合計金額": total_amount + tax,
                "ギフトフラグ": 0,
                "時間帯指定": "時間帯指定[午前中]",
                "日付指定": date_spec,
                "作業者欄": "",
                "備考": notes,
                "商品名": product_name,
                "商品コード": product_code,
                "商品価格": unit_price_bara,
                "受注数量": quantity_bara,
                "商品オプション": "",
                "出荷済フラグ": "",
                "顧客区分": 9,
                "顧客コード": "",
                "消費税率（%）": "",
                "のし": "",
                "ラッピング": "",
                "メッセージ": "",
            }
            rows.append(row)

    _archive_existing_file(csv_path)
    with open(csv_path, "w", newline="", encoding="cp932", errors="replace") as f:
        writer = csv.DictWriter(f, fieldnames=NE_HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    return csv_path


# ── クーラ（COOLA/ロジザード）用CSV ──

# 卸先マスタCSVパス
_OROSHISAKI_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "oroshisaki_master.csv")

# クーラCSVヘッダー（83カラム）
COOLA_HEADERS = [
    "ショップコード", "ショップ名", "表示受注番号", "行no",
    "配送先名", "配送先郵便番号", "配送先都道府県", "配送先住所１", "配送先住所２", "配送先住所３",
    "配送先電話番号", "配送先fax番号", "配送先eメール", "配送先法人名", "配送先所属部署・役職",
    "出荷備考", "出荷日", "配送会社id", "配送指定日", "配送時間帯",
    "顧客id", "発注日", "税種区分", "送料", "代引手数料",
    "消費税", "ポイント使用額", "ポイントステータス", "獲得ポイント", "使用ポイント",
    "合計ポイント", "請求額合計", "決済区分id", "決済方法", "配送方法",
    "領収書flg", "領収書宛名", "ギフトメッセージ", "包装1", "包装2",
    "包装料1", "包装料2", "のし", "荷送人名指定flg",
    "購入者氏名", "購入者ふりがな", "購入者郵便番号", "購入者都道府県",
    "購入者住所１", "購入者住所２", "購入者住所３",
    "購入者法人名", "購入者電話番号", "購入者eメール", "コメント",
    "品番", "品名", "形式/型番", "色id", "色名", "サイズid", "サイズ名",
    "項目選択肢", "受注数", "単価", "行備考", "出荷行備考",
    "送状備考1", "送状備考2", "出荷ステータス", "予備1", "ギフトフラグ",
    "発注時間", "優先フラグ", "クーポン金額", "納品書コメント",
    "NE配送会社ID", "認証番号", "8％消費税", "10％消費税", "8％金額合計", "10％金額合計", "消費税率",
]

_oroshisaki_cache = None


def _load_oroshisaki_master():
    """卸先マスタCSVを読み込み、卸先名→レコードのdictを返す"""
    global _oroshisaki_cache
    if _oroshisaki_cache is not None:
        return _oroshisaki_cache
    _oroshisaki_cache = {}
    if not os.path.exists(_OROSHISAKI_CSV):
        return _oroshisaki_cache
    import unicodedata
    for enc in ['shift-jis', 'cp932', 'utf-8-sig', 'utf-8']:
        try:
            with open(_OROSHISAKI_CSV, encoding=enc) as f:
                reader = csv.DictReader(f)
                for row in reader:
                    name = row.get("卸先名", "").strip()
                    if name:
                        norm = unicodedata.normalize('NFKC', name)
                        _oroshisaki_cache[name] = row
                        _oroshisaki_cache[norm] = row
            break
        except (UnicodeDecodeError, KeyError):
            continue
    return _oroshisaki_cache


def _split_address(address):
    """住所を都道府県/住所1(50文字)/住所2(15文字)/住所3に分割"""
    import re
    pref = ""
    rest = address
    m = re.match(r'(東京都|北海道|(?:京都|大阪)府|.{2,3}県)', address)
    if m:
        pref = m.group(1)
        rest = address[len(pref):]
    addr1 = rest[:50]
    addr2 = rest[50:65] if len(rest) > 50 else ""
    addr3 = rest[65:] if len(rest) > 65 else ""
    return pref, addr1, addr2, addr3


# 都道府県 → BtoB最安配送会社のマッピング
# 福山通運が安い地域は福山、それ以外はヤマト
_PREF_TO_BTOB_CARRIER_120 = {}
_PREF_TO_BTOB_CARRIER_140 = {}
# 120サイズ比較: 福山140 vs ヤマト120 → 福山が安い地域
for p in ["青森", "岩手", "秋田", "山形", "宮城", "福島",  # 北東北・南東北
          "茨城", "栃木", "群馬", "埼玉", "千葉", "東京", "神奈川", "山梨",  # 関東
          "新潟", "長野", "富山", "石川", "福井",  # 信越・北陸
          "静岡", "愛知", "岐阜", "三重",  # 東海
          "滋賀", "京都", "大阪", "兵庫", "奈良", "和歌山",  # 関西
          "鳥取", "島根", "岡山", "広島", "山口", "徳島", "香川", "愛媛", "高知"]:  # 中国・四国
    _PREF_TO_BTOB_CARRIER_120[p] = "福山通運"
# ヤマト120が安い地域: 北海道・九州・沖縄
for p in ["北海道", "福岡", "佐賀", "長崎", "熊本", "大分", "宮崎", "鹿児島", "沖縄"]:
    _PREF_TO_BTOB_CARRIER_120[p] = "ヤマト運輸"
# 140サイズ比較: 福山140 vs ヤマト140 → 同じ地域区分（九州でヤマト140が安い）
_PREF_TO_BTOB_CARRIER_140 = dict(_PREF_TO_BTOB_CARRIER_120)


def _select_shipping_method(product_name, total_cs, dest_pref):
    """商品名・CS数・都道府県から最適な配送方法を選択する。

    ルール（物流費まとめ 2026/3/9版）:
    - Snack 1-4cs: 佐川急便 60サイズ
    - Snack 5cs以上: BtoB最安（福山/ヤマト120）
    - Gummy 1cs: 佐川急便 80サイズ
    - Gummy 2cs以上: BtoB最安（福山/ヤマト120）
    - Energy: BtoB最安（福山/ヤマト140）
    - Water: BtoB最安（福山/ヤマト120）
    - その他: ヤマト運輸（デフォルト）
    """
    pn = product_name.upper() if product_name else ""
    pref_clean = (dest_pref or "").replace("都", "").replace("府", "").replace("県", "").replace("道", "")
    if pref_clean == "北海":
        pref_clean = "北海道"

    def btob_best(size="120"):
        mapping = _PREF_TO_BTOB_CARRIER_140 if size == "140" else _PREF_TO_BTOB_CARRIER_120
        carrier = mapping.get(pref_clean, "福山通運")
        if carrier == "福山通運":
            return ("福山通運", "55", "130")
        return ("ヤマト(発払い)B2v6", "20", "0812")

    if "SNACK" in pn or "サブレ" in pn or "トリュフ" in pn or "ガーリック" in pn or "ガトーショコラ" in pn:
        if total_cs < 5:
            return ("佐川急便", "10", "01")
        return btob_best("120")
    elif "GUMMY" in pn or "グミ" in pn:
        if total_cs <= 1:
            return ("佐川急便", "10", "01")
        return btob_best("120")
    elif "ENERGY" in pn or "エナジー" in pn:
        if total_cs <= 1:
            return ("佐川急便", "10", "01")
        return btob_best("140")
    elif "WATER" in pn or "ウォーター" in pn or "CERAMIDE" in pn or "セラミド" in pn:
        return btob_best("120")
    else:
        return ("ヤマト(発払い)B2v6", "20", "0812")


def _classify_product_group(product_name):
    """商品名から伝票分割グループを判定する。
    Returns: (group_key, suffix)
      - Snack 4品 → ("S", "-S") ※混載二重梱包
      - Energy    → ("E", "-E")
      - Water     → ("W", "-W")
      - Gummy     → ("G", "-G")
      - その他    → ("X", "-X")
    """
    pn = (product_name or "").upper()
    if "SNACK" in pn or "サブレ" in pn or "トリュフ" in pn or "ガーリック" in pn or "ガトーショコラ" in pn:
        return "S", "-S"
    elif "ENERGY" in pn or "エナジー" in pn:
        return "E", "-E"
    elif "WATER" in pn or "ウォーター" in pn or "CERAMIDE" in pn or "セラミド" in pn:
        return "W", "-W"
    elif "GUMMY" in pn or "グミ" in pn:
        return "G", "-G"
    else:
        return "X", "-X"


def _build_group_remarks(group_items, slip_no=""):
    """グループ内の商品から出荷備考を生成する。

    出力例（セル内改行あり）:
    発注番号00565754
    2Gummy_アテンションシールあり賞味2027/07/31 [二重梱包]
    """
    parts = []
    has_double_pack = False
    for it in group_items:
        name = it.get("master_name", it.get("ocr_name", ""))
        # 商品略称を生成（ブランド名を除去して短縮）
        short = name
        for prefix in ["2Snack ", "2Water ", "2Gummy ", "RNL 2Energy", "2Energy ", "TWO "]:
            short = short.replace(prefix, "")
        short = short.strip()
        if not short:
            short = name

        qty_cs = it.get("quantity", 0)
        try:
            qty_cs = int(float(qty_cs or 0))
        except (ValueError, TypeError):
            qty_cs = 0

        expiry = it.get("expiry_date", "")
        if expiry:
            expiry_fmt = expiry.replace("-", "/")
            parts.append(f"{short}賞味{expiry_fmt}")
        else:
            parts.append(short)

        if it.get("double_pack"):
            has_double_pack = True

    product_line = " ".join(parts)
    if has_double_pack:
        product_line += " [二重梱包]"

    # 発注番号を先頭行に、商品情報を2行目に（セル内改行）
    if slip_no:
        remarks = f"発注番号{slip_no}\n{product_line}"
    else:
        remarks = product_line
    return remarks


def results_to_coola_csv(results, pdf_name):
    """OCR結果をクーラ（COOLA/ロジザード）用CSV形式に変換して出力する。

    伝票分割ルール:
    - 商品グループ（Snack/Energy/Water/Gummy/その他）ごとに伝票を分割
    - 伝票番号にサフィックス（-S/-E/-W/-G/-X）を自動付番
    - 出荷備考にグループ単位の賞味期限・二重梱包指示を記載
    - Snack 4品は混載1伝票（二重梱包）
    """
    ensure_output_dir()
    csv_name = os.path.splitext(pdf_name)[0] + "_COOLA.csv"
    csv_path = os.path.join(OUTPUT_DIR, csv_name)

    import unicodedata
    oroshisaki = _load_oroshisaki_master()

    rows = []
    for page_result in results:
        if "error" in page_result:
            continue
        ocr = page_result["ocr_raw"]
        ddc = page_result.get("ddc_match", {})

        order_no = ocr.get("order_no", "")
        two_order_no = page_result.get("two_order_no", "")
        delivery_date = ocr.get("delivery_date", "")

        # 納品先情報
        dest_name = ddc.get("name", ocr.get("delivery_dest", ""))
        dest_postal = (ddc.get("postal", "") or "").replace("-", "")
        dest_address = ddc.get("address", "") or ""
        dest_tel = ddc.get("tel", "") or ""

        # 卸先マスタから情報取得
        oroshi = oroshisaki.get(dest_name) or oroshisaki.get(unicodedata.normalize('NFKC', dest_name)) or {}
        shop_code = "4"
        shop_name = "卸"
        payment = oroshi.get("支払方法", "その他")

        # 住所分割
        pref, addr1, addr2, addr3 = _split_address(dest_address)

        # BC列（コメント）: Amazon倉庫・自社倉庫向けは空欄、それ以外は21845
        is_amazon = dest_name.startswith("Amazon.co.jp")
        is_own_warehouse = bool(page_result.get("warehouse_direct"))
        coola_comment = "" if (is_amazon or is_own_warehouse) else "21845"

        # マッチ済み商品をグループに分類
        all_matched = [it for it in page_result.get("matched_items", []) if it.get("matched")]
        # Phase 4: COOLA は自社倉庫出荷システムなので、直送商品（ハルナ/シルビア ロット成立）は除外
        # 判定優先順位:
        #   1. item.shipping_route ("direct"/"warehouse") があればそれを尊重（ユーザー手動上書き含む）
        #   2. 無ければ output_dest と brand別CS合計から旧ロジック（10cs以上→直送、未満→自社倉庫経由）
        # ユーザーが「ロット成立だが倉庫経由にしたい」と手動切替したケースもCOOLAに含める
        from collections import defaultdict as _dd
        _cs_by_dest = _dd(int)
        for _it in all_matched:
            try:
                _cs_by_dest[_it.get("output_dest") or ""] += int(float(_it.get("quantity") or 0))
            except (TypeError, ValueError):
                pass
        def _is_direct_shipping(it):
            route = (it.get("shipping_route") or "").strip()
            if route == "warehouse":
                return False  # 倉庫経由 → COOLA対象
            if route == "direct":
                return True   # 直送 → COOLA除外
            # shipping_route 未設定の場合のフォールバック（旧ロジック）
            od = it.get("output_dest") or ""
            if od in ("シルビア", "ハルナ"):
                return _cs_by_dest[od] >= 10
            # 野村不動産 / エムズインク / POW BAR は常に直送 → COOLA除外
            if od in ("野村不動産", "エムズインク", "POW BAR"):
                return True
            return False
        matched_items = [it for it in all_matched if not _is_direct_shipping(it)]
        if not matched_items:
            continue

        groups = {}  # group_key → [items]
        for item in matched_items:
            product_name = item.get("master_name", item.get("ocr_name", ""))
            gkey, _ = _classify_product_group(product_name)
            groups.setdefault(gkey, []).append(item)

        # グループごとに伝票を生成
        suffix_map = {"S": "-S", "E": "-E", "W": "-W", "G": "-G", "X": "-X"}
        for gkey, group_items in groups.items():
            suffix = suffix_map.get(gkey, "")
            slip_no = f"{order_no}{suffix}" if len(groups) > 1 else order_no

            # グループ内の合計金額
            group_total = 0
            for item in group_items:
                irisuu = _get_irisuu(item)
                cs_p = item.get("cs_price", 0)
                qty = item.get("quantity", 0)
                try:
                    unit_p = int(float(cs_p or 0)) // irisuu if irisuu > 0 else int(float(cs_p or 0))
                    q = int(float(qty or 0)) * irisuu
                    group_total += unit_p * q
                except (ValueError, TypeError):
                    pass
            group_tax = int(group_total * 0.1)

            # 出荷備考を生成（発注番号＋商品情報をセル内改行で記載）
            remarks = _build_group_remarks(group_items, slip_no)

            # グループ内の全CS数合計（配送方法判定用）
            total_cs_in_group = 0
            for item in group_items:
                try:
                    total_cs_in_group += int(float(item.get("quantity", 0) or 0))
                except (ValueError, TypeError):
                    pass

            line_no = 0
            for item in group_items:
                line_no += 1

                qty_cs = item.get("quantity", 0)
                cs_price = item.get("cs_price", 0)
                try:
                    price = int(float(cs_price or 0))
                except (ValueError, TypeError):
                    price = 0
                try:
                    qty_cs_int = int(float(qty_cs or 0))
                except (ValueError, TypeError):
                    qty_cs_int = 0

                irisuu = _get_irisuu(item)
                quantity_bara = qty_cs_int * irisuu
                unit_price_bara = int(price / irisuu) if irisuu > 0 else price

                product_name = item.get("master_name", item.get("ocr_name", ""))
                product_code = item.get("code", "")

                # 配送方法: Snackは混載合計CSで判定、他は個別CS
                if gkey == "S":
                    shipping_method, carrier_id, time_code = _select_shipping_method(product_name, total_cs_in_group, pref)
                else:
                    shipping_method, carrier_id, time_code = _select_shipping_method(product_name, qty_cs_int, pref)

                row = {h: "" for h in COOLA_HEADERS}
                row.update({
                    "ショップコード": shop_code,
                    "ショップ名": shop_name,
                    "表示受注番号": slip_no,
                    "行no": line_no,
                    "配送先名": dest_name,
                    "配送先郵便番号": dest_postal,
                    "配送先都道府県": pref,
                    "配送先住所１": addr1,
                    "配送先住所２": addr2,
                    "配送先住所３": addr3,
                    "配送先電話番号": dest_tel,
                    "出荷備考": remarks,
                    "出荷日": "",
                    "配送会社id": carrier_id,
                    "配送指定日": f"{delivery_date} 00:00:00" if delivery_date else "",
                    "配送時間帯": time_code,
                    "顧客id": two_order_no,
                    "税種区分": "1",
                    # Phase 4: ロット割れ自社倉庫出荷時に shipping_fee_calculator で算出した送料を反映
                    # （TWO負担の場合は 0、未計算の場合も 0）
                    "送料": int(page_result.get("shipping_fee") or 0),
                    "代引手数料": 0,
                    "消費税": group_tax,
                    "ポイント使用額": 0,
                    "請求額合計": group_total + group_tax,
                    "決済区分id": "",
                    "決済方法": payment,
                    "配送方法": shipping_method,
                    "購入者氏名": dest_name,
                    "購入者郵便番号": dest_postal,
                    "購入者都道府県": pref,
                    "購入者住所１": addr1,
                    "購入者住所２": addr2,
                    "購入者住所３": addr3,
                    "購入者電話番号": dest_tel,
                    "コメント": coola_comment,
                    "品名": product_name,
                    "形式/型番": product_code,
                    "受注数": quantity_bara,
                    "単価": unit_price_bara,
                    "送状備考1": "",
                    "ギフトフラグ": 0,
                    "消費税率": 10,
                })
                rows.append(row)

    if not rows:
        return None

    _archive_existing_file(csv_path)
    with open(csv_path, "w", newline="", encoding="cp932", errors="replace") as f:
        writer = csv.DictWriter(f, fieldnames=COOLA_HEADERS, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)

    return csv_path


def results_to_excel(results, pdf_name):
    """Write OCR+matching results to Excel (.xlsx) with DDC dropdown for uncertain rows.

    - OK行: 緑背景、納品先(DDC)は固定テキスト
    - 要確認行: 黄背景、納品先(DDC)にドロップダウン（候補から選択）
    - NG行: 赤背景、納品先(DDC)にドロップダウン
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    ensure_output_dir()
    xlsx_name = os.path.splitext(pdf_name)[0] + "_result.xlsx"
    xlsx_path = os.path.join(OUTPUT_DIR, xlsx_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "受注データ"

    # カラーパレット
    C_HEADER  = "2F5496"  # ヘッダー：濃青
    C_OK      = "E2EFDA"  # OK：薄緑
    C_CAUTION = "FFF2CC"  # 要確認：薄黄
    C_NG      = "FCE4D6"  # NG：薄赤

    # 列定義: (列名, 幅)
    COLUMNS = [
        ("ページ",        5),
        ("オーダーNO",   14),
        ("納品日",        12),
        ("発注元",        22),
        ("納品先(OCR)",   28),
        ("納品先(DDC)",   30),   # ← 要確認/NG にドロップダウン
        ("DDCマッチ",      8),
        ("DDC候補1",      26),
        ("DDC候補2",      26),
        ("DDC候補3",      26),
        ("商品名(OCR)",   26),
        ("商品名(マスタ)", 26),
        ("JANコード",     14),
        ("商品コード",    10),
        ("規格",          14),
        ("配送荷姿",      10),
        ("CS単価",        10),
        ("数量(CS)",       8),
        ("金額",          10),
        ("出力先",        10),
        ("商品マッチ",     8),
        ("備考",          20),
    ]
    headers = [c[0] for c in COLUMNS]
    col_map = {h: i + 1 for i, (h, _) in enumerate(COLUMNS)}

    # ヘッダー行を書き込む
    hdr_fill = PatternFill("solid", fgColor=C_HEADER)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # 候補を格納する非表示シート（ドロップダウン参照用）
    ws_cand = wb.create_sheet("DDC候補リスト")
    ws_cand.sheet_state = "hidden"
    cand_row_idx = 0  # 候補シートの現在行

    # データ行を構築して書き込む
    ddc_col_letter = get_column_letter(col_map["納品先(DDC)"])

    for page_result in results:
        if "error" in page_result:
            continue
        ocr = page_result["ocr_raw"]
        ddc = page_result["ddc_match"]
        candidates = ddc.get("candidates", [])

        def cand_name(idx):
            return candidates[idx]["name"] if idx < len(candidates) else ""

        def cand_str(idx):
            if idx < len(candidates):
                c = candidates[idx]
                return f"{c['name']} ({c['score']:.0%})"
            return ""

        low_conf = ddc.get("low_confidence", False)
        match_status = "OK" if ddc.get("matched") else "NG"
        if ddc.get("matched") and low_conf:
            match_status = "要確認"

        # ドロップダウン用候補名リスト（空除外）
        pure_cands = [cand_name(i) for i in range(3) if cand_name(i)]

        for item in page_result["matched_items"]:
            row_data = {
                "ページ":        page_result["page"],
                "オーダーNO":    ocr.get("order_no", ""),
                "納品日":        ocr.get("delivery_date", ""),
                "発注元":        ocr.get("sender", ""),
                "納品先(OCR)":   ocr.get("delivery_dest", ""),
                "納品先(DDC)":   ddc.get("name", cand_name(0)) if ddc.get("matched") else cand_name(0),
                "DDCマッチ":     match_status,
                "DDC候補1":      cand_str(0),
                "DDC候補2":      cand_str(1),
                "DDC候補3":      cand_str(2),
                "商品名(OCR)":   item.get("ocr_name", ""),
                "商品名(マスタ)": item.get("master_name", ""),
                "JANコード":     item.get("jan", ""),
                "商品コード":    item.get("code", ""),
                "規格":          item.get("spec", ""),
                "配送荷姿":      item.get("pack", ""),
                "CS単価":        item.get("cs_price", ""),
                "数量(CS)":      item.get("quantity", ""),
                "金額":          item.get("amount", ""),
                "出力先":        item.get("output_dest", ""),
                "商品マッチ":    "OK" if item.get("matched") else "NG",
                "備考":          ocr.get("notes", ""),
            }

            # 背景色を選択
            if match_status == "OK":
                row_fill = PatternFill("solid", fgColor=C_OK)
            elif match_status == "要確認":
                row_fill = PatternFill("solid", fgColor=C_CAUTION)
            else:
                row_fill = PatternFill("solid", fgColor=C_NG)

            # 現在の書き込み行（ヘッダー=1なのでデータは2行目〜）
            data_row = ws.max_row + 1

            for col_name in headers:
                cell = ws.cell(row=data_row, column=col_map[col_name], value=row_data[col_name])
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="center")

            # 要確認/NG行にドロップダウンを追加
            if match_status in ("要確認", "NG") and pure_cands:
                cand_row_idx += 1
                # 候補を非表示シートに縦方向に書き込む
                for c_col, name in enumerate(pure_cands, 1):
                    ws_cand.cell(row=cand_row_idx, column=c_col, value=name)

                # データ検証: 候補シートの該当行を参照
                end_col = get_column_letter(len(pure_cands))
                formula = f"'DDC候補リスト'!$A${cand_row_idx}:${end_col}${cand_row_idx}"
                dv = DataValidation(
                    type="list",
                    formula1=formula,
                    showErrorMessage=False,
                    showInputMessage=True,
                    promptTitle="DDC選択",
                    prompt="正しい納品先DDCを選択してください",
                )
                dv.add(f"{ddc_col_letter}{data_row}")
                ws.add_data_validation(dv)

    # ヘッダー固定・フィルター
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    _archive_existing_file(xlsx_path)
    wb.save(xlsx_path)
    return xlsx_path


def generate_pdfs(results, pdf_name, staff_name="伊藤"):
    """Generate Sylvia and/or Haruna order PDFs from processing results"""
    ensure_output_dir()
    generated = []

    for page_result in results:
        if "error" in page_result:
            continue

        ocr = page_result["ocr_raw"]
        ddc = page_result["ddc_match"]
        sylvia_items = page_result["sylvia_items"]
        haruna_items = page_result["haruna_items"]
        remarks = page_result.get("remarks", "")
        order_no = ocr.get("order_no", "")
        dest_name = ddc.get("name", ocr.get("delivery_dest", "unknown"))
        page_num = page_result["page"]
        base_name = os.path.splitext(pdf_name)[0]

        # Sylvia PDF
        if sylvia_items:
            items_data = []
            for i in sylvia_items:
                items_data.append({
                    "jan": i.get("jan", ""),
                    "name": i["master_name"],
                    "spec": i.get("spec", ""),
                    "pack": i.get("pack", ""),
                    "unit_price": int(i.get("unit_price", 0)),
                    "cs_price": int(i.get("cs_price", 0)),
                    "quantity": i["quantity"],
                    "amount": int(i.get("amount", 0)),
                })
            order_data = {
                "order_no": order_no,
                "order_date": str(date.today()),
                "delivery_date": ocr.get("delivery_date", ""),
                "delivery_dest": dest_name,
                "postal": ddc.get("postal", ""),
                "address": ddc.get("address", ""),
                "tel": ddc.get("tel", ""),
                "fax": ddc.get("fax", ""),
                "remarks": remarks,
            }
            pdf_buf = gen_sylvia_pdf(order_data, items_data, staff_name)
            out_name = f"シルビア_{base_name}_p{page_num}.pdf"
            out_path = os.path.join(OUTPUT_DIR, out_name)
            _archive_existing_file(out_path)
            with open(out_path, "wb") as f:
                f.write(pdf_buf.read())
            generated.append(("シルビア", out_path))

        # Haruna PDF
        if haruna_items:
            total_qty = sum(i["quantity"] for i in haruna_items)
            order_data = {
                "order_no": order_no,
                "order_date": str(date.today()),
                "delivery_date": ocr.get("delivery_date", ""),
                "delivery_dest": dest_name,
                "quantity": total_qty,
                "remarks": remarks,
            }
            ddc_data = {
                "postal": ddc.get("postal", ""),
                "address": ddc.get("address", ""),
                "tel": ddc.get("tel", ""),
                "fax": ddc.get("fax", ""),
                "time": ddc.get("time", ""),
                "berse": ddc.get("berse", "無"),
                "palette": ddc.get("palette", ""),
                "jpr": ddc.get("jpr", ""),
                "method": ddc.get("method", ""),
                "notes": ddc.get("notes", ""),
            }
            pdf_buf = gen_haruna_pdf(order_data, ddc_data, staff_name)
            out_name = f"ハルナ_{base_name}_p{page_num}.pdf"
            out_path = os.path.join(OUTPUT_DIR, out_name)
            _archive_existing_file(out_path)
            with open(out_path, "wb") as f:
                f.write(pdf_buf.read())
            generated.append(("ハルナ", out_path))

        # Partner PDFs (野村不動産 / エムズインク / POW BAR) — 2026-06 追加
        # ロット条件なしで output_dest がパートナー名なら必ず発注書PDF生成
        matched_items = page_result.get("matched_items", [])
        for partner_name, (pdf_func, file_prefix) in PARTNER_PDF_GENERATORS.items():
            partner_items = [
                i for i in matched_items
                if i.get("matched") and i.get("output_dest") == partner_name
            ]
            if not partner_items:
                continue
            items_data = []
            for i in partner_items:
                items_data.append({
                    "jan": i.get("jan", ""),
                    "name": i.get("master_name", ""),
                    "spec": i.get("spec", ""),
                    "pack": i.get("pack", ""),
                    "quantity": i.get("quantity", 0),
                    # 価格情報も渡すが hide_price=True で非表示になる
                    "unit_price": int(i.get("unit_price", 0) or 0),
                    "cs_price": int(i.get("cs_price", 0) or 0),
                    "amount": int(i.get("amount", 0) or 0),
                })
            partner_order_data = {
                "order_no": order_no,
                "order_date": str(date.today()),
                "delivery_date": ocr.get("delivery_date", ""),
                "delivery_dest": dest_name,
                "postal": ddc.get("postal", ""),
                "address": ddc.get("address", ""),
                "tel": ddc.get("tel", ""),
                "fax": ddc.get("fax", ""),
                "remarks": remarks,
            }
            pdf_buf = pdf_func(partner_order_data, items_data, staff_name)
            out_name = f"{file_prefix}_{base_name}_p{page_num}.pdf"
            out_path = os.path.join(OUTPUT_DIR, out_name)
            _archive_existing_file(out_path)
            with open(out_path, "wb") as f:
                f.write(pdf_buf.read())
            generated.append((partner_name, out_path))

    return generated


def print_summary(results, csv_path, xlsx_path, generated_pdfs):
    """Print processing summary"""
    print("\n" + "=" * 60)
    print("処理結果サマリー")
    print("=" * 60)

    for page_result in results:
        page = page_result["page"]
        if "error" in page_result:
            print(f"\n  Page {page}: ERROR - {page_result['error']}")
            continue

        ocr = page_result["ocr_raw"]
        ddc = page_result["ddc_match"]
        print(f"\n  Page {page}:")
        print(f"    オーダーNO : {ocr.get('order_no', '(なし)')}")
        print(f"    納品日     : {ocr.get('delivery_date', '(なし)')}")
        print(f"    発注元     : {ocr.get('sender', '(なし)')}")
        print(f"    納品先(OCR): {ocr.get('delivery_dest', '(なし)')}")
        if ddc.get("matched"):
            score_str = ""
            if "match_score" in ddc:
                score_str = f" (類似度: {ddc['match_score']:.0%})"
            conf_str = " ⚠要確認" if ddc.get("low_confidence") else ""
            print(f"    納品先(DDC): {ddc['name']}{score_str}{conf_str}")
        else:
            print(f"    納品先(DDC): *** 未マッチ ***")

        candidates = ddc.get("candidates", [])
        if candidates:
            print(f"    DDC候補:")
            for i, c in enumerate(candidates, 1):
                print(f"      {i}. {c['name']} ({c['score']:.0%})")

        print(f"    商品:")
        for item in page_result["matched_items"]:
            if item.get("matched"):
                print(f"      OK  {item['master_name']} x {item['quantity']}CS → {item['output_dest']}")
            else:
                print(f"      NG  {item.get('ocr_name', '?')} x {item.get('quantity', '?')}CS *** 未マッチ ***")

        if page_result["sylvia_items"]:
            total = sum(i["amount"] for i in page_result["sylvia_items"])
            print(f"    → シルビア: {len(page_result['sylvia_items'])}商品, 合計 ¥{total:,.0f}")
        if page_result["haruna_items"]:
            total_cs = sum(i["quantity"] for i in page_result["haruna_items"])
            print(f"    → ハルナ: {total_cs}CS ({total_cs * 24:,}本)")

    print(f"\n  CSV出力:   {csv_path}")
    print(f"  Excel出力: {xlsx_path}")
    for dest, path in generated_pdfs:
        print(f"  {dest}PDF: {path}")
    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(description="FAX発注書 自動処理スクリプト")
    parser.add_argument("pdf_files", nargs="+", help="FAX PDFファイルのパス")
    parser.add_argument("--staff", default="伊藤", help="担当者名 (default: 伊藤)")
    parser.add_argument("--remarks", default="", help="イレギュラーリクエスト/備考")
    parser.add_argument("--model", default="claude-sonnet-4-6", help="Claude model for OCR")
    args = parser.parse_args()

    for pdf_path in args.pdf_files:
        if not os.path.exists(pdf_path):
            print(f"ERROR: ファイルが見つかりません: {pdf_path}")
            continue

        pdf_name = os.path.basename(pdf_path)
        print(f"\n{'=' * 60}")
        print(f"Processing: {pdf_name}")
        print(f"{'=' * 60}")

        # 1. OCR + matching
        results = process_fax_pdf(pdf_path, model=args.model)

        # 2. CSV + Excel output
        csv_path = results_to_csv(results, pdf_name)
        xlsx_path = results_to_excel(results, pdf_name)

        # 3. PDF generation
        generated = generate_pdfs(results, pdf_name, args.staff, args.remarks)

        # 4. Summary
        print_summary(results, csv_path, xlsx_path, generated)


if __name__ == "__main__":
    main()
